'''
Screener Excel export for the N100 Financial Intelligence Platform.

Sprint 3 Day 17 deliverable D-07: output/screener_output.xlsx with one
sheet per preset, 20 KPI columns, sorted by composite quality score
descending, and threshold cells colour-coded.

Colour coding
   green  the cell meets the preset threshold for that metric
   red    the cell fails the threshold

A red cell in a returned row is not a bug. It marks a company that was
admitted by a documented business rule rather than by the raw number:
a Financials company exempt from the D/E ceiling, or a Debt Free
company whose interest coverage is infinite.

Run with:
   python -m src.screener.export
'''

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.screener.engine import ScreenerEngine
from src.screener.universe import build_universe

OUTPUT_PATH = 'output/screener_output.xlsx'

# Identity columns shown before the KPI block.
IDENTITY_COLUMNS = ['company_id', 'company_name', 'broad_sector', 'year']

# The 20 KPI columns required by Sprint 3 Day 17.
KPI_COLUMNS = [
   'composite_quality_score',
   'sector_relative_score',
   'return_on_equity_pct',
   'return_on_capital_employed_pct',
   'net_profit_margin_pct',
   'operating_profit_margin_pct',
   'debt_to_equity',
   'interest_coverage',
   'asset_turnover',
   'free_cash_flow_cr',
   'cash_from_operations_cr',
   'revenue_cagr_5yr',
   'pat_cagr_5yr',
   'eps_cagr_5yr',
   'earnings_per_share',
   'dividend_payout_ratio_pct',
   'dividend_yield_pct',
   'pe_ratio',
   'pb_ratio',
   'market_cap_crore'
]

# Valuation metrics sourced from the simulated market_cap dataset.
SIMULATED_COLUMNS = {
   'dividend_yield_pct',
   'pe_ratio',
   'pb_ratio',
   'market_cap_crore'
}

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')

COLUMN_HEADERS = {
   'company_id': 'Ticker',
   'company_name': 'Company',
   'broad_sector': 'Sector',
   'year': 'Period',
   'composite_quality_score': 'Composite Score',
   'sector_relative_score': 'Sector Rel. Score',
   'return_on_equity_pct': 'ROE %',
   'return_on_capital_employed_pct': 'ROCE %',
   'net_profit_margin_pct': 'NPM %',
   'operating_profit_margin_pct': 'OPM %',
   'debt_to_equity': 'D/E',
   'interest_coverage': 'ICR',
   'asset_turnover': 'Asset Turnover',
   'free_cash_flow_cr': 'FCF (Cr)',
   'cash_from_operations_cr': 'CFO (Cr)',
   'revenue_cagr_5yr': 'Rev CAGR 5y %',
   'pat_cagr_5yr': 'PAT CAGR 5y %',
   'eps_cagr_5yr': 'EPS CAGR 5y %',
   'earnings_per_share': 'EPS',
   'dividend_payout_ratio_pct': 'Payout %',
   'dividend_yield_pct': 'Div Yield % *',
   'pe_ratio': 'P/E *',
   'pb_ratio': 'P/B *',
   'market_cap_crore': 'Market Cap (Cr) *'
}


def _threshold_mask(engine, result_df, preset_filters):
   # Which cells met their threshold on the raw number alone.
   masks = {}

   for column, filter_config in preset_filters.items():
      if column not in result_df.columns:
         continue

      operator_name = filter_config['operator']
      threshold = filter_config['threshold']

      if operator_name == 'between':
         mask = result_df[column].between(
            threshold['min'],
            threshold['max'],
            inclusive='both'
         )
      else:
         value = threshold.get('min', threshold.get('max'))
         mask = engine.OPERATOR_MAP[operator_name](result_df[column], value)

      masks[column] = mask.fillna(False).astype(bool)

   return masks


def _write_sheet(writer, sheet_name, result_df, threshold_masks):
   export_df = result_df[IDENTITY_COLUMNS + KPI_COLUMNS].copy()

   for column in KPI_COLUMNS:
      export_df[column] = pd.to_numeric(
         export_df[column],
         errors='coerce'
      ).round(2)

   export_df = export_df.rename(columns=COLUMN_HEADERS)
   export_df.to_excel(writer, sheet_name=sheet_name, index=False)

   worksheet = writer.sheets[sheet_name]
   all_columns = IDENTITY_COLUMNS + KPI_COLUMNS

   for column_index, column in enumerate(all_columns, start=1):
      letter = get_column_letter(column_index)
      header_cell = worksheet[f'{letter}1']
      header_cell.fill = HEADER_FILL
      header_cell.font = HEADER_FONT
      header_cell.alignment = Alignment(
         horizontal='center',
         vertical='center',
         wrap_text=True
      )

      worksheet.column_dimensions[letter].width = (
         26 if column == 'company_name' else 15
      )

      if column not in threshold_masks:
         continue

      for row_offset, passed in enumerate(threshold_masks[column], start=2):
         worksheet[f'{letter}{row_offset}'].fill = (
            PASS_FILL if passed else FAIL_FILL
         )

   worksheet.freeze_panes = 'B2'


def _write_summary_sheet(writer, summary_rows):
   summary_df = pd.DataFrame(summary_rows)
   summary_df.to_excel(writer, sheet_name='Summary', index=False)

   worksheet = writer.sheets['Summary']
   for column_index in range(1, len(summary_df.columns) + 1):
      letter = get_column_letter(column_index)
      worksheet[f'{letter}1'].fill = HEADER_FILL
      worksheet[f'{letter}1'].font = HEADER_FONT
      worksheet.column_dimensions[letter].width = 28


def export_screener_output(universe_df=None, output_path=OUTPUT_PATH):
   engine = ScreenerEngine()
   engine.load_config()

   if universe_df is None:
      universe_df = build_universe()

   universe_df = engine.add_composite_scores(universe_df)

   destination = Path(output_path)
   destination.parent.mkdir(parents=True, exist_ok=True)

   summary_rows = []

   with pd.ExcelWriter(destination, engine='openpyxl') as writer:
      for preset_name in engine.preset_names():
         result_df = engine.run_preset(preset_name, universe_df)
         preset_filters = engine.config['presets'][preset_name]['filters']

         threshold_masks = _threshold_mask(engine, result_df, preset_filters)

         # Excel sheet names are capped at 31 characters.
         sheet_name = engine.preset_label(preset_name)[:31]
         _write_sheet(writer, sheet_name, result_df, threshold_masks)

         summary_rows.append({
            'Preset': engine.preset_label(preset_name),
            'Companies Returned': len(result_df),
            'Within 5-50 Exit Range': 'Yes' if 5 <= len(result_df) <= 50
                                      else 'No',
            'Filters Applied': ', '.join(preset_filters.keys())
         })

         print(f'  {engine.preset_label(preset_name):<22} {len(result_df):>3}')

      summary_rows.append({
         'Preset': '* P/E, P/B, Dividend Yield and Market Cap are '
                   'SIMULATED datasets',
         'Companies Returned': None,
         'Within 5-50 Exit Range': None,
         'Filters Applied': None
      })

      _write_summary_sheet(writer, summary_rows)

   print(f'\nWrote {destination}')

   return destination


def main():
   print('Preset results:')
   export_screener_output()


if __name__ == '__main__':
   main()
