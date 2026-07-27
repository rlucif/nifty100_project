'''
Peer comparison Excel report for the N100 Financial Intelligence Platform.

Sprint 3 Day 20 deliverable D-09: output/peer_comparison.xlsx with one
sheet per peer group.

Each sheet carries the 10 ranked metrics as a value column plus a
percentile rank column, giving 20 metric columns in total, followed by
a peer group median summary row.

Colour coding of percentile cells
   green   >= 75th percentile
   yellow  25th to 75th percentile
   red     <= 25th percentile

The benchmark company of each peer group is highlighted in amber.

Run with:
   python -m src.reports.peer_comparison
'''

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics.peer import (
   PEER_METRICS,
   build_peer_percentiles,
   load_peer_groups
)
from src.screener.universe import build_universe

DB_PATH = 'data/nifty100.db'
OUTPUT_PATH = 'output/peer_comparison.xlsx'

UPPER_PERCENTILE_BAND = 75
LOWER_PERCENTILE_BAND = 25

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
BENCHMARK_FILL = PatternFill('solid', fgColor='FFD966')
MEDIAN_FILL = PatternFill('solid', fgColor='D9D9D9')

METRIC_LABELS = {
   'return_on_equity_pct': 'ROE %',
   'return_on_capital_employed_pct': 'ROCE %',
   'net_profit_margin_pct': 'NPM %',
   'debt_to_equity': 'D/E',
   'free_cash_flow_cr': 'FCF (Cr)',
   'pat_cagr_5yr': 'PAT CAGR 5y %',
   'revenue_cagr_5yr': 'Rev CAGR 5y %',
   'eps_cagr_5yr': 'EPS CAGR 5y %',
   'interest_coverage': 'ICR',
   'asset_turnover': 'Asset Turnover'
}


def get_connection():
   return sqlite3.connect(DB_PATH)


def build_group_sheet(group_members, percentiles_df):
   # Wide sheet: one row per company, value and rank per metric.
   wide_df = percentiles_df.pivot_table(
      index='company_id',
      columns='metric',
      values=['value', 'percentile_rank'],
      aggfunc='first'
   )

   rows = []

   for _, member in group_members.iterrows():
      company_id = member['company_id']
      row = {
         'company_id': company_id,
         'company_name': member['company_name']
      }

      for metric in PEER_METRICS:
         label = METRIC_LABELS[metric]

         try:
            value = wide_df.loc[company_id, ('value', metric)]
         except KeyError:
            value = None

         try:
            rank = wide_df.loc[company_id, ('percentile_rank', metric)]
         except KeyError:
            rank = None

         row[label] = None if pd.isna(value) else round(float(value), 2)
         row[f'{label} %ile'] = (
            None if pd.isna(rank) else round(float(rank), 1)
         )

      rows.append(row)

   sheet_df = pd.DataFrame(rows)

   # Peer group median for every value column.
   median_row = {'company_id': 'PEER MEDIAN', 'company_name': ''}
   for metric in PEER_METRICS:
      label = METRIC_LABELS[metric]
      median_row[label] = round(sheet_df[label].median(skipna=True), 2)
      median_row[f'{label} %ile'] = None

   sheet_df = pd.concat(
      [sheet_df, pd.DataFrame([median_row])],
      ignore_index=True
   )

   return sheet_df


def _percentile_fill(rank):
   if pd.isna(rank):
      return None
   if rank >= UPPER_PERCENTILE_BAND:
      return GREEN_FILL
   if rank <= LOWER_PERCENTILE_BAND:
      return RED_FILL

   return YELLOW_FILL


def _style_sheet(worksheet, sheet_df, benchmark_company):
   percentile_columns = {
      f'{METRIC_LABELS[metric]} %ile' for metric in PEER_METRICS
   }

   for column_index, column_name in enumerate(sheet_df.columns, start=1):
      letter = get_column_letter(column_index)

      header_cell = worksheet[f'{letter}1']
      header_cell.fill = HEADER_FILL
      header_cell.font = HEADER_FONT
      header_cell.alignment = Alignment(
         horizontal='center',
         vertical='center',
         wrap_text=True
      )

      worksheet.column_dimensions[letter].width = (
         26 if column_name == 'company_name' else 14
      )

      for row_offset in range(len(sheet_df)):
         excel_row = row_offset + 2
         cell = worksheet[f'{letter}{excel_row}']
         company_id = sheet_df.iloc[row_offset]['company_id']

         if company_id == 'PEER MEDIAN':
            cell.fill = MEDIAN_FILL
            cell.font = Font(bold=True)
            continue

         if company_id == benchmark_company:
            cell.fill = BENCHMARK_FILL
            cell.font = Font(bold=True)
            continue

         if column_name in percentile_columns:
            fill = _percentile_fill(sheet_df.iloc[row_offset][column_name])
            if fill is not None:
               cell.fill = fill

   worksheet.freeze_panes = 'C2'


def export_peer_comparison(output_path=OUTPUT_PATH):
   connection = get_connection()

   try:
      universe_df = build_universe(connection)
      peer_groups_df = load_peer_groups(connection)
   finally:
      connection.close()

   percentiles_df = build_peer_percentiles(universe_df, peer_groups_df)

   members_df = peer_groups_df.merge(
      universe_df[['company_id', 'company_name']],
      on='company_id',
      how='inner'
   )

   destination = Path(output_path)
   destination.parent.mkdir(parents=True, exist_ok=True)

   with pd.ExcelWriter(destination, engine='openpyxl') as writer:
      for peer_group_name, group_members in members_df.groupby(
         'peer_group_name'
      ):
         group_percentiles = percentiles_df[
            percentiles_df['peer_group_name'] == peer_group_name
         ]

         sheet_df = build_group_sheet(group_members, group_percentiles)

         # is_benchmark arrives from Excel as the string '1' or '0'.
         benchmark_flag = pd.to_numeric(
            group_members['is_benchmark'],
            errors='coerce'
         )
         benchmark_rows = group_members[benchmark_flag == 1]
         benchmark_company = (
            benchmark_rows['company_id'].iloc[0]
            if not benchmark_rows.empty else None
         )

         sheet_name = peer_group_name[:31]
         sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

         _style_sheet(writer.sheets[sheet_name], sheet_df, benchmark_company)

         print(
            f'  {peer_group_name:<22} '
            f'{len(group_members):>2} companies  '
            f'benchmark: {benchmark_company}'
         )

   print(f'\nWrote {destination}')

   return destination


def main():
   print('Peer group sheets:')
   export_peer_comparison()


if __name__ == '__main__':
   main()
