'''
Outputs
   output/cashflow_intelligence.xlsx  one row per company
   output/distress_alerts.csv         companies burning operating cash
                                      while raising financing
   output/pattern_changes.csv         capital allocation pattern moves
'''

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics.cashflow_kpis import (
   calculate_average_cfo_quality_score,
   calculate_capex_intensity,
   calculate_cfo_quality_score,
   calculate_deleveraging_flag,
   calculate_distress_flag,
   get_capex_intensity_label,
   get_cfo_quality_label
)
from src.analytics.periods import (
   add_period_columns,
   deduplicate_company_years
)

DB_PATH = 'data/nifty100.db'
INTELLIGENCE_PATH = 'output/cashflow_intelligence.xlsx'
DISTRESS_PATH = 'output/distress_alerts.csv'
PATTERN_CHANGES_PATH = 'output/pattern_changes.csv'
CAPITAL_ALLOCATION_CSV = 'output/capital_allocation.csv'

CFO_QUALITY_WINDOW_YEARS = 5
FCF_CAGR_WINDOW_YEARS = 5

OUTPUT_COLUMNS = [
   'company_id',
   'company_name',
   'sector',
   'year',
   'cfo_quality_score',
   'cfo_quality_label',
   'capex_intensity_pct',
   'capex_label',
   'fcf_cagr_5yr',
   'fcf_conversion_pct',
   'distress_flag',
   'deleveraging_flag',
   'capital_allocation_label'
]

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
ALERT_FILL = PatternFill('solid', fgColor='FFC7CE')
GOOD_FILL = PatternFill('solid', fgColor='C6EFCE')

LABEL_FILLS = {
   'High Quality Earnings': GOOD_FILL,
   'Accrual Risk': ALERT_FILL,
   'Asset Light': GOOD_FILL,
   'Capital Intensive': PatternFill('solid', fgColor='FFEB9C')
}


def get_connection():
   return sqlite3.connect(DB_PATH)


def load_frames(connection):
   cashflow_df = deduplicate_company_years(
      pd.read_sql('SELECT * FROM cashflow', connection)
   )
   profit_df = deduplicate_company_years(
      pd.read_sql(
         'SELECT company_id, year, sales, operating_profit, net_profit '
         'FROM profitandloss',
         connection
      )
   )
   balance_df = deduplicate_company_years(
      pd.read_sql(
         'SELECT company_id, year, borrowings FROM balancesheet',
         connection
      )
   )
   sectors_df = pd.read_sql(
      'SELECT company_id, broad_sector FROM sectors',
      connection
   )
   companies_df = pd.read_sql(
      'SELECT id AS company_id, company_name FROM companies',
      connection
   )

   history = (
      cashflow_df
      .merge(profit_df, on=['company_id', 'year'], how='left')
      .merge(balance_df, on=['company_id', 'year'], how='left')
      .merge(sectors_df, on='company_id', how='left')
      .merge(companies_df, on='company_id', how='left')
   )

   history = add_period_columns(history)
   history = history[history['period_sort_key'] > 0]

   return history.sort_values(['company_id', 'period_sort_key'])


def _cagr(start_value, end_value, years):
   if pd.isna(start_value) or pd.isna(end_value):
      return None
   if start_value <= 0 or end_value < 0:
      return None

   return round(((end_value / start_value) ** (1 / years) - 1) * 100, 2)


def build_intelligence(history, allocation_df):
   records = []

   latest_patterns = _latest_patterns(allocation_df)

   for company_id, group in history.groupby('company_id'):
      group = group.sort_values('period_sort_key')
      latest = group.iloc[-1]

      recent = group.tail(CFO_QUALITY_WINDOW_YEARS)
      yearly_scores = [
         calculate_cfo_quality_score(operating_activity, net_profit)
         for operating_activity, net_profit in zip(
            recent['operating_activity'],
            recent['net_profit']
         )
         if pd.notna(operating_activity) and pd.notna(net_profit)
      ]
      average_score = calculate_average_cfo_quality_score(yearly_scores)

      capex_intensity = (
         calculate_capex_intensity(
            latest['investing_activity'],
            latest['sales']
         )
         if pd.notna(latest['investing_activity'])
         and pd.notna(latest['sales']) else None
      )

      free_cash_flow = (
         latest['operating_activity'] + latest['investing_activity']
         if pd.notna(latest['operating_activity'])
         and pd.notna(latest['investing_activity']) else None
      )

      fcf_conversion = (
         round(free_cash_flow / latest['operating_profit'] * 100, 2)
         if free_cash_flow is not None
         and pd.notna(latest['operating_profit'])
         and latest['operating_profit'] != 0 else None
      )

      fcf_series = (
         group['operating_activity'] + group['investing_activity']
      ).reset_index(drop=True)
      fcf_cagr = (
         _cagr(
            fcf_series.iloc[-(FCF_CAGR_WINDOW_YEARS + 1)],
            fcf_series.iloc[-1],
            FCF_CAGR_WINDOW_YEARS
         )
         if len(fcf_series) > FCF_CAGR_WINDOW_YEARS else None
      )

      previous_borrowings = (
         group['borrowings'].iloc[-2] if len(group) > 1 else None
      )

      records.append({
         'company_id': company_id,
         'company_name': latest.get('company_name'),
         'sector': latest.get('broad_sector'),
         'year': latest.get('year'),
         'cfo_quality_score': average_score,
         'cfo_quality_label': get_cfo_quality_label(average_score),
         'capex_intensity_pct': capex_intensity,
         'capex_label': get_capex_intensity_label(capex_intensity),
         'fcf_cagr_5yr': fcf_cagr,
         'fcf_conversion_pct': fcf_conversion,
         'distress_flag': calculate_distress_flag(
            latest['operating_activity'],
            latest['financing_activity']
         ),
         'deleveraging_flag': calculate_deleveraging_flag(
            latest['financing_activity'],
            latest['borrowings'],
            previous_borrowings
         ),
         'capital_allocation_label': latest_patterns.get(company_id),
         'operating_activity': latest['operating_activity'],
         'financing_activity': latest['financing_activity'],
         'net_profit': latest['net_profit']
      })

   return pd.DataFrame(records).sort_values('company_id')


def _latest_patterns(allocation_df):
   if allocation_df.empty:
      return {}

   patterns = add_period_columns(allocation_df)
   patterns = patterns[patterns['period_sort_key'] > 0]
   patterns = patterns.sort_values(['company_id', 'period_sort_key'])
   latest = patterns.groupby('company_id').tail(1)

   return dict(zip(latest['company_id'], latest['pattern_label']))


def build_pattern_changes(allocation_df):
   # Companies whose capital allocation pattern changed year over year.
   if allocation_df.empty:
      return pd.DataFrame(
         columns=[
            'company_id', 'from_year', 'to_year',
            'from_pattern', 'to_pattern'
         ]
      )

   patterns = add_period_columns(allocation_df)
   patterns = patterns[patterns['period_sort_key'] > 0]
   patterns = patterns.sort_values(['company_id', 'period_sort_key'])

   grouped = patterns.groupby('company_id')
   patterns['previous_pattern'] = grouped['pattern_label'].shift(1)
   patterns['previous_year'] = grouped['year'].shift(1)

   changes = patterns[
      patterns['previous_pattern'].notna()
      & (patterns['previous_pattern'] != patterns['pattern_label'])
   ]

   return changes[[
      'company_id',
      'previous_year',
      'year',
      'previous_pattern',
      'pattern_label'
   ]].rename(columns={
      'previous_year': 'from_year',
      'year': 'to_year',
      'previous_pattern': 'from_pattern',
      'pattern_label': 'to_pattern'
   })


def verify_capital_allocation(allocation_df, expected_companies):
   present = set(allocation_df['company_id'])
   missing = sorted(expected_companies - present)

   unknown = allocation_df[
      allocation_df['pattern_label'] == 'Unknown Pattern'
   ]

   return {
      'rows': len(allocation_df),
      'companies': len(present),
      'missing_companies': missing,
      'unknown_pattern_rows': len(unknown)
   }


def _write_sheet(writer, sheet_name, frame):
   frame.to_excel(writer, sheet_name=sheet_name, index=False)
   worksheet = writer.sheets[sheet_name]

   for column_index, column_name in enumerate(frame.columns, start=1):
      letter = get_column_letter(column_index)
      header = worksheet[f'{letter}1']
      header.fill = HEADER_FILL
      header.font = HEADER_FONT
      header.alignment = Alignment(
         horizontal='center', vertical='center', wrap_text=True
      )
      worksheet.column_dimensions[letter].width = (
         26 if column_name in ('company_name', 'capital_allocation_label')
         else 16
      )

      for row_offset in range(len(frame)):
         cell = worksheet[f'{letter}{row_offset + 2}']
         value = frame.iloc[row_offset][column_name]

         if column_name in ('cfo_quality_label', 'capex_label'):
            fill = LABEL_FILLS.get(value)
            if fill:
               cell.fill = fill

         if column_name == 'distress_flag' and value:
            cell.fill = ALERT_FILL
         if column_name == 'deleveraging_flag' and value:
            cell.fill = GOOD_FILL

   worksheet.freeze_panes = 'B2'


def export_cashflow_intelligence(connection=None):
   owns_connection = connection is None
   if owns_connection:
      connection = get_connection()

   try:
      history = load_frames(connection)
      expected = set(
         pd.read_sql(
            'SELECT DISTINCT company_id FROM financial_ratios',
            connection
         )['company_id']
      )
   finally:
      if owns_connection:
         connection.close()

   # The cashflow table carries 100 tickers, but the platform universe is
   # the 92 companies that survived the Sprint 2 statement joins. Without
   # this filter the report would run to 100 rows and the extra 8 would
   # appear as unclassified, since they have no ratio or pattern data.
   history = history[history['company_id'].isin(expected)]

   allocation_path = Path(CAPITAL_ALLOCATION_CSV)
   allocation_df = (
      pd.read_csv(allocation_path) if allocation_path.exists()
      else pd.DataFrame(columns=['company_id', 'year', 'pattern_label'])
   )

   intelligence_df = build_intelligence(history, allocation_df)
   changes_df = build_pattern_changes(allocation_df)
   verification = verify_capital_allocation(allocation_df, expected)

   export_df = intelligence_df[OUTPUT_COLUMNS].copy()

   distribution = (
      export_df['capital_allocation_label']
      .fillna('Not classified')
      .value_counts()
      .rename_axis('Pattern')
      .reset_index(name='Companies')
   )

   destination = Path(INTELLIGENCE_PATH)
   destination.parent.mkdir(parents=True, exist_ok=True)

   with pd.ExcelWriter(destination, engine='openpyxl') as writer:
      _write_sheet(writer, 'Cash Flow Intelligence', export_df)
      _write_sheet(writer, 'Pattern Distribution', distribution)

      notes = pd.DataFrame({'Note': [
         'CFO quality score is CFO/PAT averaged over the last '
         f'{CFO_QUALITY_WINDOW_YEARS} years.',
         'High Quality above 1.0, Moderate 0.5 to 1.0, '
         'Accrual Risk below 0.5.',
         'CapEx intensity is abs(investing activity) / sales x 100. '
         'Asset Light below 3%, Moderate 3-8%, Capital Intensive above 8%.',
         'Distress flag: operating cash flow negative while financing '
         'cash flow is positive in the latest year.',
         'Caution on the distress flag: for banks and NBFCs this is the '
         'normal signature of loan book growth, not distress. '
         'output/distress_alerts.csv marks those rows in the '
         'structurally_normal_for_sector column.',
         'Deleveraging flag: financing cash flow negative and the '
         'borrowings balance fell year over year.',
         'Operating profit is EBITDA by construction: the source P&L '
         'reports it before depreciation, verified against the identity '
         'operating profit + other income - interest - depreciation = '
         'profit before tax.'
      ]})
      _write_sheet(writer, 'Notes', notes)

   distress_df = intelligence_df[intelligence_df['distress_flag']][[
      'company_id',
      'company_name',
      'sector',
      'year',
      'operating_activity',
      'financing_activity',
      'net_profit',
      'cfo_quality_score',
      'capital_allocation_label'
   ]].rename(columns={
      'operating_activity': 'cfo_cr',
      'financing_activity': 'cff_cr',
      'net_profit': 'net_profit_cr'
   }).copy()

   distress_df['structurally_normal_for_sector'] = [
      sector == 'Financials' and pd.notna(net_profit) and net_profit > 0
      for sector, net_profit in zip(
         distress_df['sector'],
         distress_df['net_profit_cr']
      )
   ]

   distress_df['interpretation'] = [
      'Lending growth pattern, profitable - not a distress indicator'
      if normal else
      'Operating cash outflow funded by financing - review'
      for normal in distress_df['structurally_normal_for_sector']
   ]

   distress_df.to_csv(Path(DISTRESS_PATH), index=False)
   changes_df.to_csv(Path(PATTERN_CHANGES_PATH), index=False)

   print(f'Wrote {destination} ({len(export_df)} companies)')
   print(f'Wrote {DISTRESS_PATH} ({len(distress_df)} distress alerts)')
   print(f'Wrote {PATTERN_CHANGES_PATH} ({len(changes_df)} pattern changes)')
   print()
   print('Capital allocation completeness check:')
   print(f'  rows                : {verification["rows"]}')
   print(f'  companies           : {verification["companies"]}')
   print(f'  missing companies   : {verification["missing_companies"]}')
   print(f'  unknown pattern rows: {verification["unknown_pattern_rows"]}')
   print()
   print('Latest-year pattern distribution:')
   print(distribution.to_string(index=False))
   print()
   print('Labels:')
   print(export_df['cfo_quality_label'].value_counts(dropna=False).to_string())
   print(export_df['capex_label'].value_counts(dropna=False).to_string())
   print(f'deleveraging: {int(export_df["deleveraging_flag"].sum())}')

   return export_df, distress_df, changes_df


def main():
   export_cashflow_intelligence()


if __name__ == '__main__':
   main()
